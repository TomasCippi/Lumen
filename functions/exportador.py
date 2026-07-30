import os
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────────────────────────────────

NOMBRE_SUBCARPETA_INDIVIDUALES = "Excels individuales"
LARGO_MAXIMO_NOMBRE_HOJA = 31  # límite que impone Excel para el nombre de una hoja

# Color para destacar la columna "Precio Total" en todos los Excel
COLOR_DESTACADO_TOTAL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

# Colores para diferenciar de un vistazo las columnas de Descuento y Aumento
COLOR_DESCUENTO = PatternFill(start_color="F8C9C9", end_color="F8C9C9", fill_type="solid")  # rojo clarito
COLOR_AUMENTO = PatternFill(start_color="C9E8C9", end_color="C9E8C9", fill_type="solid")    # verde clarito

# Borde fino y gris para toda la tabla
BORDE_FINO = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


# ─────────────────────────────────────────────────────────────────────────
# Helpers internos
# ─────────────────────────────────────────────────────────────────────────

def _sanear_nombre_hoja(nombre: str) -> str:
    """Limpia caracteres inválidos para nombres de hoja de Excel y recorta el largo."""
    nombre_limpio = re.sub(r'[\\/?*\[\]:]', "", nombre).strip()
    if not nombre_limpio:
        nombre_limpio = "Proveedor"
    return nombre_limpio[:LARGO_MAXIMO_NOMBRE_HOJA]


def _parsear_porcentajes_encadenados(texto: str) -> list:
    """
    Convierte un string tipo "29-3-3" en [29.0, 3.0, 3.0].
    Si el texto viene vacío o no se puede parsear, devuelve [0.0]
    (una sola columna en 0%, para que siempre exista al menos una).
    """
    if not texto:
        return [0.0]

    partes = [p for p in texto.split("-") if p.strip() != ""]
    if not partes:
        return [0.0]

    try:
        return [float(p) for p in partes]
    except ValueError:
        return [0.0]


def _construir_hoja_vendedor(
    hoja,
    productos: list,
    nombre_proveedor: str,
    incluir_familia: bool,
    porcentaje_vendedor: float,
    valor_dolar: float,
    texto_descuento: str = "",
    texto_aumento: str = "",
) -> None:
    """
    Arma la hoja de Excel completa para un proveedor: título, encabezados,
    columnas dinámicas de descuento/aumento encadenados, filas de productos
    con fórmulas, bordes y anchos de columna.
    """
    descuentos = _parsear_porcentajes_encadenados(texto_descuento)
    aumentos = _parsear_porcentajes_encadenados(texto_aumento)

    cantidad_descuentos = len(descuentos)
    cantidad_aumentos = len(aumentos)

    # --- Definición de columnas -------------------------------------------
    encabezados = ["Código", "Descripción"]
    if incluir_familia:
        encabezados.append("Familia")
    encabezados.append("Precio")

    # Columnas de Descuento encadenadas (una o más)
    for i in range(cantidad_descuentos):
        encabezados.append("Descuento" if cantidad_descuentos == 1 else f"Descuento {i + 1}")

    # Columnas de Aumento encadenadas (una o más)
    for i in range(cantidad_aumentos):
        encabezados.append("Aumento" if cantidad_aumentos == 1 else f"Aumento {i + 1}")

    encabezados.append("Precio Vendedor")
    if valor_dolar:
        encabezados.append("Dólar")
    encabezados.append("Precio Total")  # siempre presente, al final

    total_columnas = len(encabezados)

    col_precio = encabezados.index("Precio") + 1
    primera_col_descuento = col_precio + 1
    primera_col_aumento = primera_col_descuento + cantidad_descuentos
    col_precio_vendedor = primera_col_aumento + cantidad_aumentos
    col_dolar = col_precio_vendedor + 1 if valor_dolar else None
    col_precio_total = total_columnas  # siempre la última

    letra_precio = get_column_letter(col_precio)
    letra_precio_vendedor = get_column_letter(col_precio_vendedor)
    letra_dolar = get_column_letter(col_dolar) if valor_dolar else None

    fila_encabezados = 2
    formato_precio_origen = '"US$ "#,##0.00' if valor_dolar else '"$"#,##0.00'

    # --- Fila 1: título con el nombre del proveedor ------------------------
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columnas)
    celda_titulo = hoja.cell(row=1, column=1, value=nombre_proveedor)
    celda_titulo.font = Font(size=26, bold=True)
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
    hoja.row_dimensions[1].height = 40

    # --- Fila 2: encabezados de texto (base) --------------------------------
    for i, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=fila_encabezados, column=i, value=texto)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

        if primera_col_descuento <= i < primera_col_descuento + cantidad_descuentos:
            celda.fill = COLOR_DESCUENTO
        elif primera_col_aumento <= i < primera_col_aumento + cantidad_aumentos:
            celda.fill = COLOR_AUMENTO

    # --- Encabezados "especiales": números editables en vez de texto -------

    # Descuentos: cada columna tiene su propio % editable (default 0%)
    letras_descuento = []
    for i, valor_pct in enumerate(descuentos):
        col = primera_col_descuento + i
        letra = get_column_letter(col)
        letras_descuento.append(letra)
        celda = hoja.cell(row=fila_encabezados, column=col, value=valor_pct / 100)
        celda.number_format = '"DESCUENTO "0.00%'
        celda.font = Font(bold=True)
        celda.fill = COLOR_DESCUENTO

    # Aumentos: misma lógica que los descuentos
    letras_aumento = []
    for i, valor_pct in enumerate(aumentos):
        col = primera_col_aumento + i
        letra = get_column_letter(col)
        letras_aumento.append(letra)
        celda = hoja.cell(row=fila_encabezados, column=col, value=valor_pct / 100)
        celda.number_format = '"AUMENTO "0.00%'
        celda.font = Font(bold=True)
        celda.fill = COLOR_AUMENTO

    # % del vendedor (editable)
    celda_pct_vendedor = hoja.cell(row=fila_encabezados, column=col_precio_vendedor, value=porcentaje_vendedor / 100)
    celda_pct_vendedor.number_format = "0.00%"
    celda_pct_vendedor.font = Font(bold=True)

    # Valor del dólar (editable), se muestra como "DÓLAR: 1400"
    if valor_dolar:
        celda_dolar = hoja.cell(row=fila_encabezados, column=col_dolar, value=valor_dolar)
        celda_dolar.number_format = '"DÓLAR: "0'
        celda_dolar.font = Font(bold=True)

    # Encabezado de "Precio Total" (siempre presente, con color destacado)
    celda_total_header = hoja.cell(row=fila_encabezados, column=col_precio_total, value="Precio Total")
    celda_total_header.font = Font(bold=True)
    celda_total_header.alignment = Alignment(horizontal="center")
    celda_total_header.fill = COLOR_DESTACADO_TOTAL

    # --- Filas de productos --------------------------------------------------
    fila_excel = fila_encabezados + 1
    for producto in productos:
        columna = 1
        hoja.cell(row=fila_excel, column=columna, value=producto["codigo"])
        columna += 1
        hoja.cell(row=fila_excel, column=columna, value=producto["producto"])
        columna += 1

        if incluir_familia:
            hoja.cell(row=fila_excel, column=columna, value=producto.get("familia", ""))
            columna += 1

        # Precio de origen (compra)
        hoja.cell(row=fila_excel, column=col_precio, value=producto["precio"]).number_format = formato_precio_origen

        # "letra_base" es la columna que toma cada fórmula como punto de
        # partida; arranca en Precio y se va actualizando a medida que se
        # encadenan Descuento -> Aumento -> Precio Vendedor.
        letra_base = letra_precio

        # Descuentos encadenados: cada uno toma el resultado del anterior
        for i, letra_pct in enumerate(letras_descuento):
            col_actual = primera_col_descuento + i
            hoja.cell(
                row=fila_excel, column=col_actual,
                value=f"=ROUND({letra_base}{fila_excel}*(1-${letra_pct}${fila_encabezados}),2)",
            ).number_format = formato_precio_origen
            letra_base = get_column_letter(col_actual)

        # Aumentos encadenados: misma lógica, pero sumando el %
        for i, letra_pct in enumerate(letras_aumento):
            col_actual = primera_col_aumento + i
            hoja.cell(
                row=fila_excel, column=col_actual,
                value=f"=ROUND({letra_base}{fila_excel}*(1+${letra_pct}${fila_encabezados}),2)",
            ).number_format = formato_precio_origen
            letra_base = get_column_letter(col_actual)

        # Precio Vendedor: toma el resultado final de Descuento/Aumento
        hoja.cell(
            row=fila_excel, column=col_precio_vendedor,
            value=f"=ROUND({letra_base}{fila_excel}*${letra_precio_vendedor}${fila_encabezados},2)",
        ).number_format = formato_precio_origen

        if valor_dolar:
            # Dólar -> pesos: Precio Vendedor (USD) * valor del dólar
            hoja.cell(
                row=fila_excel, column=col_dolar,
                value=f"=ROUND({letra_precio_vendedor}{fila_excel}*${letra_dolar}${fila_encabezados},2)",
            ).number_format = '"$"#,##0.00'
            formula_total = f"=ROUND({letra_dolar}{fila_excel},2)"
        else:
            # Sin dólar: Precio Total es directamente el Precio Vendedor
            formula_total = f"=ROUND({letra_precio_vendedor}{fila_excel},2)"

        celda_total = hoja.cell(row=fila_excel, column=col_precio_total, value=formula_total)
        celda_total.number_format = '"$"#,##0.00'
        celda_total.fill = COLOR_DESTACADO_TOTAL

        fila_excel += 1

    # --- Bordes finos en toda la tabla (encabezado + filas de datos) --------
    ultima_fila_con_datos = fila_excel - 1
    for fila in range(fila_encabezados, ultima_fila_con_datos + 1):
        for col in range(1, total_columnas + 1):
            hoja.cell(row=fila, column=col).border = BORDE_FINO

    # --- Ancho de columnas ----------------------------------------------------
    anchos = {1: 14, 2: 45, col_precio: 14, col_precio_vendedor: 16, col_precio_total: 14}
    if incluir_familia:
        anchos[3] = 30
    if valor_dolar:
        anchos[col_dolar] = 16
    for col in range(primera_col_descuento, primera_col_descuento + cantidad_descuentos):
        anchos[col] = 14
    for col in range(primera_col_aumento, primera_col_aumento + cantidad_aumentos):
        anchos[col] = 14
    for col, ancho in anchos.items():
        hoja.column_dimensions[get_column_letter(col)].width = ancho


def _guardar_excel_individual(
    productos, nombre_proveedor, ruta_carpeta_destino,
    incluir_familia, porcentaje_vendedor, valor_dolar,
    texto_descuento, texto_aumento,
) -> str:
    """Genera el Excel individual (con fecha en el nombre) y devuelve su ruta."""
    carpeta_individuales = os.path.join(ruta_carpeta_destino, NOMBRE_SUBCARPETA_INDIVIDUALES)
    os.makedirs(carpeta_individuales, exist_ok=True)

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Vendedor"

    _construir_hoja_vendedor(
        hoja, productos, nombre_proveedor, incluir_familia, porcentaje_vendedor,
        valor_dolar, texto_descuento, texto_aumento,
    )

    marca_tiempo = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nombre_archivo = f"{nombre_proveedor} - {marca_tiempo}.xlsx"
    ruta_final = os.path.join(carpeta_individuales, nombre_archivo)

    try:
        libro.save(ruta_final)
    except PermissionError:
        raise PermissionError(f"El archivo '{nombre_archivo}' está abierto. Cerralo y volvé a intentar.")

    return ruta_final


def _guardar_en_base_de_datos(
    productos, nombre_proveedor, ruta_carpeta_destino,
    incluir_familia, porcentaje_vendedor, valor_dolar,
    texto_descuento, texto_aumento,
) -> str:
    """
    Agrega (o reemplaza) la hoja de este proveedor dentro del Excel
    "Base de Datos Vendedor {año}.xlsx". Devuelve la ruta de ese archivo.
    """
    anio_actual = datetime.now().year
    nombre_archivo_bd = f"Base de Datos Vendedor {anio_actual}.xlsx"
    ruta_bd = os.path.join(ruta_carpeta_destino, nombre_archivo_bd)

    if os.path.exists(ruta_bd):
        libro = openpyxl.load_workbook(ruta_bd)
    else:
        libro = openpyxl.Workbook()
        libro.remove(libro.active)  # arrancamos con un libro sin hojas

    nombre_hoja = _sanear_nombre_hoja(nombre_proveedor)

    if nombre_hoja in libro.sheetnames:
        libro.remove(libro[nombre_hoja])

    hoja = libro.create_sheet(title=nombre_hoja)

    _construir_hoja_vendedor(
        hoja, productos, nombre_proveedor, incluir_familia, porcentaje_vendedor,
        valor_dolar, texto_descuento, texto_aumento,
    )

    try:
        libro.save(ruta_bd)
    except PermissionError:
        raise PermissionError(f"El archivo '{nombre_archivo_bd}' está abierto. Cerralo y volvé a intentar.")

    return ruta_bd


# ─────────────────────────────────────────────────────────────────────────
# API pública
# ─────────────────────────────────────────────────────────────────────────

def exportar_vendedor(
    productos: list,
    nombre_proveedor: str,
    ruta_carpeta_destino: str,
    incluir_familia: bool,
    porcentaje_vendedor: float = 181.5,
    valor_dolar: float = None,
    texto_descuento: str = "",
    texto_aumento: str = "",
) -> dict:
    """
    Exporta los datos del proveedor de dos maneras:
        1. Un Excel individual (con fecha), dentro de "Excels individuales".
        2. Una hoja (con el nombre del proveedor) dentro de la base de
           datos anual.

    texto_descuento / texto_aumento: strings tal como los escribió el
    usuario, ej. "29-3-3" (varios porcentajes encadenados) o "" (sin
    valor; igual se crea una columna en 0%).

    Devuelve {"individual": "...", "base_de_datos": "..."}.
    """
    os.makedirs(ruta_carpeta_destino, exist_ok=True)

    ruta_individual = _guardar_excel_individual(
        productos, nombre_proveedor, ruta_carpeta_destino,
        incluir_familia, porcentaje_vendedor, valor_dolar,
        texto_descuento, texto_aumento,
    )

    ruta_bd = _guardar_en_base_de_datos(
        productos, nombre_proveedor, ruta_carpeta_destino,
        incluir_familia, porcentaje_vendedor, valor_dolar,
        texto_descuento, texto_aumento,
    )

    return {"individual": ruta_individual, "base_de_datos": ruta_bd}


def exportar_stock_facil(
    productos: list,
    nombre_proveedor: str,
    ruta_carpeta_destino: str,
    incluir_familia: bool,
    precio_total_por_producto: dict,
) -> str:
    """
    Exporta al formato XLS que usa Stock Fácil.

    precio_total_por_producto: dict {codigo: precio_total_calculado} con
    el precio final (ya con descuento/aumento/vendedor aplicado).

    Devuelve la ruta del archivo generado.
    """
    import xlwt

    def _limpiar_texto(valor: str) -> str:
        return str(valor).replace(",", ".").replace("'", '"')

    libro = xlwt.Workbook(encoding="utf-8")
    hoja = libro.add_sheet("stock")

    encabezados = [
        "codigo", "descripcion", "marca", "cantidad", "precio1", "precio2",
        "precio3", "precio4", "familia", "proveedor", "mayor1", "cantidad1", "iva",
    ]
    for col, texto in enumerate(encabezados):
        hoja.write(0, col, texto)

    for fila_idx, producto in enumerate(productos, start=1):
        codigo = _limpiar_texto(producto["codigo"])
        descripcion = _limpiar_texto(producto["producto"])
        marca = _limpiar_texto(nombre_proveedor)
        precio1 = round(float(producto["precio"]), 2)
        precio2 = round(float(precio_total_por_producto.get(producto["codigo"], producto["precio"])), 2)
        familia = _limpiar_texto(producto.get("familia", "") or nombre_proveedor) if incluir_familia else _limpiar_texto(nombre_proveedor)
        proveedor = _limpiar_texto(nombre_proveedor)

        hoja.write(fila_idx, 0, codigo)
        hoja.write(fila_idx, 1, descripcion)
        hoja.write(fila_idx, 2, marca)
        hoja.write(fila_idx, 3, 0)   # cantidad
        hoja.write(fila_idx, 4, precio1)
        hoja.write(fila_idx, 5, precio2)
        hoja.write(fila_idx, 6, 0)   # precio3
        hoja.write(fila_idx, 7, 0)   # precio4
        hoja.write(fila_idx, 8, familia)
        hoja.write(fila_idx, 9, proveedor)
        hoja.write(fila_idx, 10, 0)  # mayor1
        hoja.write(fila_idx, 11, 0)  # cantidad1
        hoja.write(fila_idx, 12, 0)  # iva

    os.makedirs(ruta_carpeta_destino, exist_ok=True)
    marca_tiempo = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nombre_archivo = f"{nombre_proveedor} - {marca_tiempo} - StockFacil.xls"
    ruta_final = os.path.join(ruta_carpeta_destino, nombre_archivo)

    try:
        libro.save(ruta_final)
    except PermissionError:
        raise PermissionError(f"El archivo '{nombre_archivo}' está abierto. Cerralo y volvé a intentar.")

    return ruta_final