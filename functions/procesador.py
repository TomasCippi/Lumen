"""
procesador.py

Procesamiento del Excel: extracción de productos.
----------------------------------------------------
Recorremos el archivo UNA sola vez, fila por fila, en orden (con
iter_rows), porque acceder a celdas sueltas al azar (ej. hoja["C50"])
es muy lento en archivos grandes cuando se usa modo de solo lectura.

No asumimos que los productos arrancan en una fila fija: puede haber
texto, instrucciones o encabezados antes de la tabla real. Por eso
bajamos fila por fila hasta encontrar el PRIMER número válido en la
columna de Precio; ahí arranca la lectura de productos.

Lógica de corte de la tabla:
    - Si una fila no tiene precio válido, se revisan las 3 filas
      siguientes. Si NINGUNA de esas 3 tiene precio, se asume que ahí
      terminó la tabla de productos y se corta la lectura.
    - Si alguna de esas 3 sí tiene precio, la fila actual (sin precio)
      simplemente se ignora y se sigue leyendo.

Errores por fila:
    - Si una fila SÍ tiene precio válido, pero falta Código o
      Producto, esa fila se guarda en la lista de errores y NO se
      incluye en el resultado final (el usuario la completa a mano
      después, directamente en el Excel).
"""

import re

import openpyxl

FILAS_A_REVISAR_ANTES_DE_CORTAR = 3


# ─────────────────────────────────────────────────────────────────────────
# Helpers de limpieza de valores
# ─────────────────────────────────────────────────────────────────────────

def _limpiar_precio(valor):
    """
    Intenta convertir el valor de una celda a un número.
    Soporta texto con símbolos, ej: "$1.500,00" -> 1500.00
    Devuelve None si no se pudo interpretar como número.
    """
    if valor is None:
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return None

    texto = re.sub(r"[^\d.,-]", "", texto)
    if not texto:
        return None

    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        return None


def _texto(valor) -> str:
    """Convierte el valor de una celda a string prolijo (sin espacios extra, sin None)."""
    if valor is None:
        return ""
    return str(valor).strip()


# ─────────────────────────────────────────────────────────────────────────
# Extracción de productos
# ─────────────────────────────────────────────────────────────────────────

def procesar_excel(
    ruta_archivo: str,
    columna_codigo: str,
    columna_producto: str,
    columna_precio: str,
    columna_familia: str = None,
) -> dict:
    """
    Recorre el Excel UNA sola vez (secuencialmente) y devuelve:
        {
            "productos": [
                {"fila": 6, "codigo": "...", "producto": "...",
                 "precio": 1500.0, "familia": "..." (si aplica)},
                ...
            ],
            "errores": [
                {"fila": 20, "motivo": "Falta el código"},
                ...
            ],
        }
    """
    productos = []
    errores = []

    libro = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
    hoja = libro.active

    idx_codigo = openpyxl.utils.column_index_from_string(columna_codigo) if columna_codigo else None
    idx_producto = openpyxl.utils.column_index_from_string(columna_producto) if columna_producto else None
    idx_precio = openpyxl.utils.column_index_from_string(columna_precio)
    idx_familia = openpyxl.utils.column_index_from_string(columna_familia) if columna_familia else None

    def valor_en(fila_valores, indice):
        """Extrae el valor de una tupla de fila dado un índice de columna (1-based)."""
        if indice is None:
            return None
        if indice - 1 >= len(fila_valores):
            return None
        return fila_valores[indice - 1]

    # Guardamos todas las filas en memoria como lista de valores, en una
    # sola pasada secuencial (esto es rápido incluso con miles de filas).
    todas_las_filas = [fila for fila in hoja.iter_rows(values_only=True)]
    libro.close()

    total_filas = len(todas_las_filas)

    def tiene_precio(numero_fila_1indexed):
        """numero_fila_1indexed es el número de fila real del Excel (1, 2, 3...)."""
        indice_lista = numero_fila_1indexed - 1
        if indice_lista < 0 or indice_lista >= total_filas:
            return False
        valores = todas_las_filas[indice_lista]
        return _limpiar_precio(valor_en(valores, idx_precio)) is not None

    # 1) Buscar dónde arranca la tabla real: primera fila con precio válido
    fila_actual = 1
    while fila_actual <= total_filas and not tiene_precio(fila_actual):
        fila_actual += 1

    if fila_actual > total_filas:
        return {"productos": [], "errores": []}  # nunca hubo un precio

    # 2) Recorrer desde ahí aplicando la lógica de productos/errores/corte
    while fila_actual <= total_filas:
        valores = todas_las_filas[fila_actual - 1]
        precio = _limpiar_precio(valor_en(valores, idx_precio))

        if precio is None:
            hay_precio_mas_adelante = any(
                tiene_precio(fila_actual + offset)
                for offset in range(1, FILAS_A_REVISAR_ANTES_DE_CORTAR + 1)
            )
            if not hay_precio_mas_adelante:
                break  # se terminó la tabla de productos

            fila_actual += 1
            continue

        codigo = _texto(valor_en(valores, idx_codigo))
        producto = _texto(valor_en(valores, idx_producto))
        familia = _texto(valor_en(valores, idx_familia)) if idx_familia else ""

        if not codigo or not producto:
            faltantes = []
            if not codigo:
                faltantes.append("código")
            if not producto:
                faltantes.append("producto")

            errores.append({
                "fila": fila_actual,
                "motivo": f"Falta el {' y el '.join(faltantes)}",
            })
            fila_actual += 1
            continue

        item = {
            "fila": fila_actual,
            "codigo": codigo,
            "producto": producto,
            "precio": precio,
        }
        if idx_familia:
            item["familia"] = familia

        productos.append(item)
        fila_actual += 1

    return {"productos": productos, "errores": errores}