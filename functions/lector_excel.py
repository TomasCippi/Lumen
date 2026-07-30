"""
lector.py

Detección automática de columnas de un Excel.
------------------------------------------------
Al elegir un archivo, buscamos la fila de ENCABEZADOS para autocompletar
los dropdowns de Código, Producto, Precio y Familia. Es solo una ayuda:
si no encuentra algo, lo deja vacío y el usuario elige la letra a mano.

Importante: NO asumimos que los encabezados están en la fila 1, porque
muchos Excel reales tienen texto/instrucciones antes (como en este caso,
donde los encabezados están en la fila 5). Por eso recorremos las
primeras filas hasta encontrar la que tiene cara de encabezados.

La búsqueda de palabras clave es APROXIMADA: "Descripción del producto"
cuenta como coincidencia de "producto", por ejemplo.
"""

import os
import openpyxl

# ─────────────────────────────────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────────────────────────────────

# Palabras clave por campo. Se busca en minúscula y sin importar
# si el encabezado tiene más texto alrededor.
PALABRAS_CLAVE = {
    "codigo": ["codigo", "código", "cod"],
    "producto": ["producto", "descripcion", "descripción", "detalle"],
    "precio": ["precio neto", "precio"],   # "precio neto" se chequea primero
    "familia": ["familia", "rubro", "categoria", "categoría"],
}

# Cuántas filas iniciales revisamos como candidatas a ser la de encabezados
FILAS_A_REVISAR_PARA_ENCABEZADOS = 15

# ─────────────────────────────────────────────────────────────────────────
# Detección de columnas
# ─────────────────────────────────────────────────────────────────────────

def _letra_columna(numero_columna: int) -> str:
    """Convierte un número de columna (1, 2, 3...) a letra de Excel (A, B, C...)."""
    return openpyxl.utils.get_column_letter(numero_columna)


def detectar_columnas(ruta_archivo: str) -> dict:
    """
    Busca la fila de encabezados dentro de las primeras filas del Excel
    y devuelve un dict con la letra encontrada para cada campo, por ej:
        {"codigo": "A", "producto": "C", "precio": "D", "familia": "B"}
    Si no encuentra alguno, ese campo queda en None.
    """
    resultado = {"codigo": None, "producto": None, "precio": None, "familia": None}

    try:
        libro = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
        hoja = libro.active
    except Exception:
        return resultado  # si el archivo no se puede abrir, devolvemos todo vacío

    for fila in hoja.iter_rows(min_row=1, max_row=FILAS_A_REVISAR_PARA_ENCABEZADOS):
        encontrados_en_esta_fila = {}

        for celda in fila:
            texto_encabezado = str(celda.value or "").strip().lower()
            if not texto_encabezado:
                continue

            for campo, palabras in PALABRAS_CLAVE.items():
                if campo in encontrados_en_esta_fila:
                    continue
                for palabra in palabras:
                    if palabra in texto_encabezado:
                        encontrados_en_esta_fila[campo] = _letra_columna(celda.column)
                        break

        # Si esta fila tiene al menos código, producto o precio, la damos
        # por buena como fila de encabezados y no seguimos mirando más abajo.
        if {"codigo", "producto", "precio"} & encontrados_en_esta_fila.keys():
            resultado.update(encontrados_en_esta_fila)
            break

    libro.close()
    return resultado